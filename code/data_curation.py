"""
Session 1: Data Curation for KG-Informed Bayesian Prior PoS Project
====================================================================
Generates:
  - data/elo_krd_patients.csv        (patient-level Elo-KRd data)
  - data/trials_data.csv             (30-40 NDMM trial arms)
  - data/drug_target_map.csv         (drug-class-target mapping)
  - data/master_simulated_patients.csv
  - data/manhattan_simulated_patients.csv
"""

import os
import warnings
import numpy as np
import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_FILE = os.path.join(BASE_DIR, "EloKRd cfDNA PET raw data_6_2025.xlsx")
os.makedirs(DATA_DIR, exist_ok=True)


# ===========================================================================
# 1. Parse Elo-KRd Excel
# ===========================================================================

def parse_elo_krd(excel_path: str = EXCEL_FILE) -> pd.DataFrame:
    """Parse enrolled-patient sheet from the Elo-KRd trial Excel file.

    Returns one row per patient with cleaned clinical variables.

    Column mapping (1-indexed in Excel, header in row 2):
      1  Subject ID
      47 Age at Enrollment
      48 Sex
      54 ISS staging
      55 R-ISS staging
      60 High-risk Cyto (1=Yes, 0=No)
      61-69 Individual cytogenetics
      72 Double-hit myeloma
      75 Cycle 4 Response
      79 Cycle 8 Response
      85 Cycle 12 Response
      92 Best Overall Response through Cutoff
      93 MRD < 10^-5 at any time (1=Yes)
      94 MRD < 10^-6 at any time (1=Yes)
      95 sCR and/or MRD <10^-5 at any time (1=Yes)
      105 Met primary endpoint window (1=Yes)
      106 Cycle 8: sCR and/or MRD-negativity (1=Yes)
      119 PFS status (1=prog or death)
      120 OS status
      122 PFS (months)
      126 OS through Cut-off Date (months)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Enrolled Patients"]

    col_map = {
        "subject_id": 1,
        "age": 47,
        "sex": 48,
        "iss": 54,
        "r_iss": 55,
        "high_risk_cyto": 60,
        "t_4_14": 61,
        "t_14_16": 62,
        "t_14_20": 63,
        "del_17p": 64,
        "p53_mut": 65,
        "gain_1q21": 66,
        "gain_1q21_3copies": 67,
        "gain_1q21_amp": 68,
        "del_1p": 69,
        "double_hit": 72,
        "response_c4": 75,
        "response_c8": 79,
        "response_c12": 85,
        "best_response": 92,
        "mrd_neg_1e5_anytime": 93,
        "mrd_neg_1e6_anytime": 94,
        "scr_or_mrd_neg_anytime": 95,
        "met_primary_endpoint_window": 105,
        "primary_endpoint": 106,
        "pfs_status": 119,
        "os_status": 120,
        "pfs_months": 122,
        "os_months": 126,
    }

    rows = []
    for r in range(3, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is None:
            continue
        row = {}
        for col_name, col_idx in col_map.items():
            row[col_name] = ws.cell(r, col_idx).value
        rows.append(row)
    wb.close()

    df = pd.DataFrame(rows)

    # Clean types
    df["subject_id"] = df["subject_id"].astype(str)
    for c in ["iss", "r_iss"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")  # nullable int
    for c in [
        "high_risk_cyto", "t_4_14", "t_14_16", "t_14_20", "del_17p",
        "p53_mut", "gain_1q21", "gain_1q21_3copies", "gain_1q21_amp",
        "del_1p", "double_hit",
        "mrd_neg_1e5_anytime", "mrd_neg_1e6_anytime",
        "scr_or_mrd_neg_anytime", "met_primary_endpoint_window",
        "primary_endpoint", "pfs_status", "os_status",
    ]:
        df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
    df["age"] = pd.to_numeric(df["age"], errors="coerce")
    df["pfs_months"] = pd.to_numeric(df["pfs_months"], errors="coerce")
    df["os_months"] = pd.to_numeric(df["os_months"], errors="coerce")

    return df


# ===========================================================================
# 2. Build trials_data.csv
# ===========================================================================

# Hard-coded from published papers / conference abstracts.
# Each entry is one trial arm. Multiple arms from same trial have distinct rows.
# p_hat = y / n (MRD-negativity rate at the specified threshold and timepoint).
# Where papers report % we convert; y = round(p_hat * n).

MANUAL_TRIALS = [
    # --- Elo-KRd (our case study) ---
    # Published aggregate: 26/45 = 0.578 (use published, not file-based)
    {
        "trial_id": "elokrd",
        "trial_name": "Elo-KRd",
        "nct_number": "NCT02969837",
        "phase": 2,
        "arm_label": "Elo-KRd",
        "n": 45,
        "y": 26,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "cycle8",
        "asct_required": 0,
        "drugs": "elotuzumab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 63,
        "frac_high_risk": 0.40,
        "frac_iss3": 0.27,
        "population": "mixed",
        "source": "Derman et al. JAMA Oncol 2022",
        "pubmed_id": "35862034",
        "notes": "File has n=30 rows (data cut); published n=45, 26 responders. Using published aggregate.",
    },
    # --- PERSEUS ---
    {
        "trial_id": "perseus_dvrd",
        "trial_name": "PERSEUS",
        "nct_number": "NCT03710603",
        "phase": 3,
        "arm_label": "D-VRd",
        "n": 355,
        "y": 266,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "daratumumab,bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 59,
        "frac_high_risk": 0.19,
        "frac_iss3": 0.15,
        "population": "TE",
        "source": "Sonneveld et al. NEJM 2024",
        "pubmed_id": "38507751",
        "notes": "MRD-neg 75.2% at 12 months",
    },
    {
        "trial_id": "perseus_vrd",
        "trial_name": "PERSEUS",
        "nct_number": "NCT03710603",
        "phase": 3,
        "arm_label": "VRd",
        "n": 354,
        "y": 170,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 59,
        "frac_high_risk": 0.17,
        "frac_iss3": 0.16,
        "population": "TE",
        "source": "Sonneveld et al. NEJM 2024",
        "pubmed_id": "38507751",
        "notes": "MRD-neg 47.5%",
    },
    # --- CEPHEUS ---
    {
        "trial_id": "cepheus_dvrd",
        "trial_name": "CEPHEUS",
        "nct_number": "NCT03652363",
        "phase": 3,
        "arm_label": "D-VRd",
        "n": 197,
        "y": 120,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "daratumumab,bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 73,
        "frac_high_risk": 0.16,
        "frac_iss3": 0.18,
        "population": "TIE",
        "source": "Facon et al. Nat Med 2025",
        "pubmed_id": "",
        "notes": "MRD-neg 60.9%",
    },
    {
        "trial_id": "cepheus_vrd",
        "trial_name": "CEPHEUS",
        "nct_number": "NCT03652363",
        "phase": 3,
        "arm_label": "VRd",
        "n": 198,
        "y": 77,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 73,
        "frac_high_risk": 0.14,
        "frac_iss3": 0.17,
        "population": "TIE",
        "source": "Facon et al. Nat Med 2025",
        "pubmed_id": "",
        "notes": "MRD-neg 39.4%",
    },
    # --- GRIFFIN ---
    {
        "trial_id": "griffin_dvrd",
        "trial_name": "GRIFFIN",
        "nct_number": "NCT02874742",
        "phase": 2,
        "arm_label": "D-VRd",
        "n": 104,
        "y": 67,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "daratumumab,bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 59,
        "frac_high_risk": 0.16,
        "frac_iss3": 0.12,
        "population": "TE",
        "source": "Voorhees et al. Lancet Haematol 2023",
        "pubmed_id": "37708911",
        "notes": "MRD-neg 64.4% end of consolidation",
    },
    {
        "trial_id": "griffin_vrd",
        "trial_name": "GRIFFIN",
        "nct_number": "NCT02874742",
        "phase": 2,
        "arm_label": "VRd",
        "n": 103,
        "y": 31,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 58,
        "frac_high_risk": 0.17,
        "frac_iss3": 0.11,
        "population": "TE",
        "source": "Voorhees et al. Lancet Haematol 2023",
        "pubmed_id": "37708911",
        "notes": "MRD-neg 30.1%",
    },
    # --- CASSIOPEIA ---
    {
        "trial_id": "cassiopeia_dvtd",
        "trial_name": "CASSIOPEIA",
        "nct_number": "NCT02541383",
        "phase": 3,
        "arm_label": "D-VTd",
        "n": 543,
        "y": 347,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "day100-post-ASCT",
        "asct_required": 1,
        "drugs": "daratumumab,bortezomib,thalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 58,
        "frac_high_risk": 0.15,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Moreau et al. Lancet 2019",
        "pubmed_id": "31097401",
        "notes": "MRD-neg 64% day100 post-ASCT",
    },
    {
        "trial_id": "cassiopeia_vtd",
        "trial_name": "CASSIOPEIA",
        "nct_number": "NCT02541383",
        "phase": 3,
        "arm_label": "VTd",
        "n": 542,
        "y": 238,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "day100-post-ASCT",
        "asct_required": 1,
        "drugs": "bortezomib,thalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 58,
        "frac_high_risk": 0.15,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Moreau et al. Lancet 2019",
        "pubmed_id": "31097401",
        "notes": "MRD-neg 44%",
    },
    # --- MAIA ---
    {
        "trial_id": "maia_drd",
        "trial_name": "MAIA",
        "nct_number": "NCT02252172",
        "phase": 3,
        "arm_label": "D-Rd",
        "n": 368,
        "y": 118,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "any-time",
        "asct_required": 0,
        "drugs": "daratumumab,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 73,
        "frac_high_risk": 0.15,
        "frac_iss3": 0.17,
        "population": "TIE",
        "source": "Facon et al. NEJM 2019; updated Lancet Oncol 2021",
        "pubmed_id": "31141632",
        "notes": "MRD-neg 32.1% at primary analysis",
    },
    {
        "trial_id": "maia_rd",
        "trial_name": "MAIA",
        "nct_number": "NCT02252172",
        "phase": 3,
        "arm_label": "Rd",
        "n": 369,
        "y": 41,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "any-time",
        "asct_required": 0,
        "drugs": "lenalidomide,dexamethasone",
        "triplet_or_quad": "doublet",
        "median_age": 73,
        "frac_high_risk": 0.15,
        "frac_iss3": 0.20,
        "population": "TIE",
        "source": "Facon et al. NEJM 2019",
        "pubmed_id": "31141632",
        "notes": "MRD-neg 11.1%",
    },
    # --- ALCYONE ---
    {
        "trial_id": "alcyone_dvmp",
        "trial_name": "ALCYONE",
        "nct_number": "NCT02195479",
        "phase": 3,
        "arm_label": "D-VMP",
        "n": 350,
        "y": 98,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "daratumumab,bortezomib,melphalan,prednisone",
        "triplet_or_quad": "quad",
        "median_age": 71,
        "frac_high_risk": 0.14,
        "frac_iss3": 0.17,
        "population": "TIE",
        "source": "Mateos et al. NEJM 2018",
        "pubmed_id": "29231133",
        "notes": "MRD-neg 28.0%",
    },
    {
        "trial_id": "alcyone_vmp",
        "trial_name": "ALCYONE",
        "nct_number": "NCT02195479",
        "phase": 3,
        "arm_label": "VMP",
        "n": 356,
        "y": 25,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "bortezomib,melphalan,prednisone",
        "triplet_or_quad": "triplet",
        "median_age": 71,
        "frac_high_risk": 0.14,
        "frac_iss3": 0.19,
        "population": "TIE",
        "source": "Mateos et al. NEJM 2018",
        "pubmed_id": "29231133",
        "notes": "MRD-neg 7.0%",
    },
    # --- GMMG-HD7 ---
    {
        "trial_id": "gmmghd7_isavrd",
        "trial_name": "GMMG-HD7",
        "nct_number": "NCT03617731",
        "phase": 3,
        "arm_label": "Isa-VRd",
        "n": 330,
        "y": 165,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 1,
        "drugs": "isatuximab,bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 58,
        "frac_high_risk": 0.15,
        "frac_iss3": 0.16,
        "population": "TE",
        "source": "Goldschmidt et al. Lancet Haematol 2022",
        "pubmed_id": "36370730",
        "notes": "MRD-neg 50.1% post-induction",
    },
    {
        "trial_id": "gmmghd7_vrd",
        "trial_name": "GMMG-HD7",
        "nct_number": "NCT03617731",
        "phase": 3,
        "arm_label": "VRd",
        "n": 330,
        "y": 119,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 1,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 57,
        "frac_high_risk": 0.17,
        "frac_iss3": 0.16,
        "population": "TE",
        "source": "Goldschmidt et al. Lancet Haematol 2022",
        "pubmed_id": "36370730",
        "notes": "MRD-neg 35.6%",
    },
    # --- BENEFIT ---
    {
        "trial_id": "benefit_isavrd",
        "trial_name": "BENEFIT",
        "nct_number": "NCT04751877",
        "phase": 3,
        "arm_label": "Isa-VRd",
        "n": 135,
        "y": 72,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "isatuximab,bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 72,
        "frac_high_risk": 0.17,
        "frac_iss3": 0.18,
        "population": "TIE",
        "source": "Facon et al. ASCO 2024",
        "pubmed_id": "",
        "notes": "MRD-neg 53.3%",
    },
    {
        "trial_id": "benefit_isard",
        "trial_name": "BENEFIT",
        "nct_number": "NCT04751877",
        "phase": 3,
        "arm_label": "Isa-Rd",
        "n": 135,
        "y": 35,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "isatuximab,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 72,
        "frac_high_risk": 0.16,
        "frac_iss3": 0.19,
        "population": "TIE",
        "source": "Facon et al. ASCO 2024",
        "pubmed_id": "",
        "notes": "MRD-neg 26.1%",
    },
    # --- GMMG-CONCEPT (Isa-KRd) ---
    {
        "trial_id": "gmmg_concept_isakrd",
        "trial_name": "GMMG-CONCEPT",
        "nct_number": "NCT03104842",
        "phase": 2,
        "arm_label": "Isa-KRd",
        "n": 99,
        "y": 67,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "isatuximab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 56,
        "frac_high_risk": 1.0,
        "frac_iss3": 0.30,
        "population": "TE",
        "source": "Leypoldt et al. JCO 2024",
        "pubmed_id": "37753960",
        "notes": "MRD-neg 67.7% post-consolidation in high-risk TE patients; ITT n=99",
    },
    # --- FORTE ---
    {
        "trial_id": "forte_krd_asct",
        "trial_name": "FORTE",
        "nct_number": "NCT02203643",
        "phase": 2,
        "arm_label": "KRd-ASCT-KRd",
        "n": 158,
        "y": 92,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 56,
        "frac_high_risk": 0.18,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Gay et al. Lancet Oncol 2021",
        "pubmed_id": "34171281",
        "notes": "MRD-neg 58% post-consolidation",
    },
    {
        "trial_id": "forte_krd12",
        "trial_name": "FORTE",
        "nct_number": "NCT02203643",
        "phase": 2,
        "arm_label": "KRd12",
        "n": 158,
        "y": 88,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 56,
        "frac_high_risk": 0.18,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Gay et al. Lancet Oncol 2021",
        "pubmed_id": "34171281",
        "notes": "MRD-neg 56% no-ASCT arm",
    },
    # --- MASTER ---
    {
        "trial_id": "master",
        "trial_name": "MASTER",
        "nct_number": "NCT03224507",
        "phase": 2,
        "arm_label": "Dara-KRd",
        "n": 123,
        "y": 87,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "daratumumab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 57,
        "frac_high_risk": 0.32,
        "frac_iss3": 0.13,
        "population": "TE",
        "source": "Costa et al. JCO 2022",
        "pubmed_id": "35271306",
        "notes": "MRD-neg 71%",
    },
    # --- MANHATTAN ---
    {
        "trial_id": "manhattan",
        "trial_name": "MANHATTAN",
        "nct_number": "NCT04096066",
        "phase": 2,
        "arm_label": "Dara-KRd",
        "n": 41,
        "y": 29,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "daratumumab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 58,
        "frac_high_risk": 0.41,
        "frac_iss3": 0.15,
        "population": "mixed",
        "source": "Landgren et al. JAMA Oncol 2021",
        "pubmed_id": "33636137",
        "notes": "MRD-neg 71%",
    },
    # --- ADVANCE ---
    {
        "trial_id": "advance_dakrd",
        "trial_name": "ADVANCE",
        "nct_number": "NCT04268498",
        "phase": 3,
        "arm_label": "DKRd",
        "n": 148,
        "y": 87,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "daratumumab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 57,
        "frac_high_risk": 0.16,
        "frac_iss3": 0.12,
        "population": "TE",
        "source": "Landgren et al. JCO 2025; 43(suppl 16): 7503",
        "pubmed_id": "",
        "notes": "MRD-neg 58.8% (ITT). ASCO 2025 abstract.",
    },
    {
        "trial_id": "advance_krd",
        "trial_name": "ADVANCE",
        "nct_number": "NCT04268498",
        "phase": 3,
        "arm_label": "KRd",
        "n": 139,
        "y": 46,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 57,
        "frac_high_risk": 0.17,
        "frac_iss3": 0.11,
        "population": "TE",
        "source": "Landgren et al. JCO 2025; 43(suppl 16): 7503",
        "pubmed_id": "",
        "notes": "MRD-neg 33.1% (ITT). ASCO 2025 abstract.",
    },
    # --- MIDAS ---
    {
        "trial_id": "midas_isakrd",
        "trial_name": "MIDAS",
        "nct_number": "NCT04934475",
        "phase": 3,
        "arm_label": "Isa-KRd (induction)",
        "n": 791,
        "y": 498,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "isatuximab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 58,
        "frac_high_risk": 0.19,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Perrot et al. Blood 2025; 146(1): 52-61",
        "pubmed_id": "",
        "notes": "MRD-neg 63% unconditional at 10^-5 after 6-cycle Isa-KRd induction (ITT). Randomization to Isa-KRd vs ASCT occurs post-MRD.",
    },
    # --- ELOQUENT-1: DROPPED (MRD-negativity never reported as endpoint) ---
    # --- IFM2009 ---
    {
        "trial_id": "ifm2009_vrd_asct",
        "trial_name": "IFM2009",
        "nct_number": "NCT01191060",
        "phase": 3,
        "arm_label": "VRd+ASCT",
        "n": 350,
        "y": 220,
        "mrd_threshold": "1e-4",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 59,
        "frac_high_risk": 0.19,
        "frac_iss3": 0.18,
        "population": "TE",
        "source": "Attal et al. NEJM 2017",
        "pubmed_id": "28379796",
        "notes": "MRD-neg 62.9% at 10^-4 (flow). 10^-5 NGS rates lower.",
    },
    {
        "trial_id": "ifm2009_vrd_noasct",
        "trial_name": "IFM2009",
        "nct_number": "NCT01191060",
        "phase": 3,
        "arm_label": "VRd-no-ASCT",
        "n": 350,
        "y": 171,
        "mrd_threshold": "1e-4",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 0,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 59,
        "frac_high_risk": 0.19,
        "frac_iss3": 0.18,
        "population": "TE",
        "source": "Attal et al. NEJM 2017",
        "pubmed_id": "28379796",
        "notes": "MRD-neg 49% at 10^-4 (flow)",
    },
    # --- DETERMINATION ---
    {
        "trial_id": "determination_vrd_asct",
        "trial_name": "DETERMINATION",
        "nct_number": "NCT01208662",
        "phase": 3,
        "arm_label": "VRd+ASCT",
        "n": 365,
        "y": 199,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 56,
        "frac_high_risk": 0.18,
        "frac_iss3": 0.12,
        "population": "TE",
        "source": "Richardson et al. NEJM 2022",
        "pubmed_id": "36355058",
        "notes": "MRD-neg 54.5% (substudy rate extrapolated to ITT n=365)",
    },
    {
        "trial_id": "determination_vrd",
        "trial_name": "DETERMINATION",
        "nct_number": "NCT01208662",
        "phase": 3,
        "arm_label": "VRd-alone",
        "n": 357,
        "y": 142,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 0,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 56,
        "frac_high_risk": 0.18,
        "frac_iss3": 0.12,
        "population": "TE",
        "source": "Richardson et al. NEJM 2022",
        "pubmed_id": "36355058",
        "notes": "MRD-neg 39.8% (substudy rate extrapolated to ITT n=357)",
    },
    # --- SWOG-S1211: DROPPED (MRD-negativity never reported) ---
    # --- ISKIA ---
    {
        "trial_id": "iskia_isakrd",
        "trial_name": "IsKia",
        "nct_number": "NCT04483739",
        "phase": 3,
        "arm_label": "Isa-KRd",
        "n": 151,
        "y": 119,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "isatuximab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 58,
        "frac_high_risk": 0.22,
        "frac_iss3": 0.13,
        "population": "TE",
        "source": "Gay et al. Blood 2023; 142(Suppl 1): 4",
        "pubmed_id": "",
        "notes": "MRD-neg 77% post-consolidation",
    },
    {
        "trial_id": "iskia_krd",
        "trial_name": "IsKia",
        "nct_number": "NCT04483739",
        "phase": 3,
        "arm_label": "KRd",
        "n": 151,
        "y": 95,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-consolidation",
        "asct_required": 1,
        "drugs": "carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 57,
        "frac_high_risk": 0.21,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Gay et al. Blood 2023; 142(Suppl 1): 4",
        "pubmed_id": "",
        "notes": "MRD-neg 67% post-consolidation. Note: 10^-5 rate was 63% in ITT; using evaluable.",
    },
    # --- ENDURANCE ---
    {
        "trial_id": "endurance_krd",
        "trial_name": "ENDURANCE",
        "nct_number": "NCT01863550",
        "phase": 3,
        "arm_label": "KRd",
        "n": 526,
        "y": 54,
        "mrd_threshold": "1e-4",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 63,
        "frac_high_risk": 0.0,
        "frac_iss3": 0.10,
        "population": "mixed",
        "source": "Kumar et al. Lancet Oncol 2020",
        "pubmed_id": "32866432",
        "notes": "MRD-neg 10.3%. Standard-risk only (excluded high-risk). Flow cytometry, threshold unspecified.",
    },
    {
        "trial_id": "endurance_vrd",
        "trial_name": "ENDURANCE",
        "nct_number": "NCT01863550",
        "phase": 3,
        "arm_label": "VRd",
        "n": 527,
        "y": 38,
        "mrd_threshold": "1e-4",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "bortezomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "triplet",
        "median_age": 63,
        "frac_high_risk": 0.0,
        "frac_iss3": 0.10,
        "population": "mixed",
        "source": "Kumar et al. Lancet Oncol 2020",
        "pubmed_id": "32866432",
        "notes": "MRD-neg 7.2%. Standard-risk only. Flow cytometry, threshold unspecified.",
    },
    # --- Derman-DKRd (formerly listed as COBRA — corrected trial identity) ---
    {
        "trial_id": "derman_dakrd",
        "trial_name": "Derman-DKRd",
        "nct_number": "NCT03500445",
        "phase": 2,
        "arm_label": "Dara-KRd",
        "n": 42,
        "y": 26,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 0,
        "drugs": "daratumumab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 69,
        "frac_high_risk": 0.24,
        "frac_iss3": 0.10,
        "population": "mixed",
        "source": "Derman et al. Blood Cancer J 2024; 14(1): 87",
        "pubmed_id": "38811560",
        "notes": "MRD-neg 61.9% (20/32 evaluable; ITT n=42). UChicago single-center.",
    },
    # --- IFM2018-04 ---
    {
        "trial_id": "ifm201804_dakrd",
        "trial_name": "IFM2018-04",
        "nct_number": "NCT03606577",
        "phase": 2,
        "arm_label": "Dara-KRd",
        "n": 50,
        "y": 31,
        "mrd_threshold": "1e-5",
        "mrd_timepoint": "post-induction",
        "asct_required": 1,
        "drugs": "daratumumab,carfilzomib,lenalidomide,dexamethasone",
        "triplet_or_quad": "quad",
        "median_age": 57,
        "frac_high_risk": 0.30,
        "frac_iss3": 0.14,
        "population": "TE",
        "source": "Touzeau et al. Blood 2024; 143(20): 2029-2036",
        "pubmed_id": "38394666",
        "notes": "MRD-neg 62% post-induction at 10^-5 (23/37 evaluable; ITT n=50). French high-risk enriched.",
    },
]


def build_trials_data() -> pd.DataFrame:
    """Build trials_data.csv from manually curated trial data."""
    df = pd.DataFrame(MANUAL_TRIALS)

    # Compute p_hat
    df["p_hat"] = df["y"] / df["n"]

    # Number of drugs
    df["n_drugs"] = df["drugs"].apply(lambda x: len(x.split(",")))

    # Binary drug flags
    all_drugs_in_data = set()
    for d in df["drugs"]:
        all_drugs_in_data.update(d.split(","))

    for drug in sorted(all_drugs_in_data):
        col_name = f"contains_{drug.replace('-', '_')}"
        df[col_name] = df["drugs"].apply(lambda x, d=drug: int(d in x.split(",")))

    return df


# ===========================================================================
# 3. Drug-Target Map
# ===========================================================================

DRUG_TARGET_DATA = [
    ("elotuzumab", "mAb", "SLAMF7", "Anti-SLAMF7 monoclonal antibody; activates NK cells"),
    ("carfilzomib", "PI", "PSMB5 (20S proteasome)", "Irreversible proteasome inhibitor"),
    ("bortezomib", "PI", "PSMB5 (20S proteasome)", "Reversible proteasome inhibitor"),
    ("lenalidomide", "IMiD", "CRBN (cereblon)", "Immunomodulatory drug; cereblon E3 ligase modulator"),
    ("pomalidomide", "IMiD", "CRBN (cereblon)", "Immunomodulatory drug; next-gen cereblon modulator"),
    ("thalidomide", "IMiD", "CRBN (cereblon)", "Immunomodulatory drug; first-gen cereblon binder"),
    ("dexamethasone", "steroid", "GR (glucocorticoid receptor)", "Corticosteroid; anti-inflammatory and pro-apoptotic"),
    ("prednisone", "steroid", "GR (glucocorticoid receptor)", "Corticosteroid"),
    ("daratumumab", "mAb", "CD38", "Anti-CD38 monoclonal antibody"),
    ("isatuximab", "mAb", "CD38", "Anti-CD38 monoclonal antibody (binds distinct epitope from daratumumab)"),
    ("melphalan", "alkylator", "DNA", "Nitrogen mustard alkylating agent"),
]


def build_drug_target_map() -> pd.DataFrame:
    """Build drug_target_map.csv."""
    df = pd.DataFrame(
        DRUG_TARGET_DATA,
        columns=["drug_name", "drug_class", "targets", "mechanism"],
    )
    return df


# ===========================================================================
# 4. Simulate Patient-Level Data
# ===========================================================================

def simulate_patient_responses(
    n: int,
    p_hat: float,
    frac_high_risk: float,
    p_hat_hr: float,
    p_hat_sr: float,
    frac_iss3: float = 0.15,
    median_age: float = 57.0,
    seed: int = 42,
) -> pd.DataFrame:
    """Simulate patient-level binary response data.

    Each patient gets:
      - high_risk ~ Bernoulli(frac_high_risk)
      - iss3 ~ Bernoulli(frac_iss3)
      - age ~ Normal(median_age, 8) clipped to [30, 90]
      - response ~ Bernoulli(p_hr if high_risk else p_sr)

    Validates that the marginal response rate is within ±5% of p_hat.
    """
    rng = np.random.default_rng(seed)
    high_risk = rng.binomial(1, frac_high_risk, size=n)
    iss3 = rng.binomial(1, frac_iss3, size=n)
    age = np.clip(rng.normal(median_age, 8, size=n), 30, 90).round(1)

    prob = np.where(high_risk == 1, p_hat_hr, p_hat_sr)
    response = rng.binomial(1, prob)

    df = pd.DataFrame({
        "patient_id": np.arange(1, n + 1),
        "age": age,
        "high_risk": high_risk,
        "iss3": iss3,
        "response": response,
    })

    observed_rate = df["response"].mean()
    if abs(observed_rate - p_hat) > 0.05:
        warnings.warn(
            f"Simulated response rate {observed_rate:.3f} differs from target "
            f"{p_hat:.3f} by more than 5%. Consider adjusting p_hat_hr/p_hat_sr or seed."
        )

    return df


# ===========================================================================
# 5. Main: Generate All Data Files
# ===========================================================================

def main():
    print("=" * 60)
    print("Session 1: Data Curation")
    print("=" * 60)

    # --- 1. Elo-KRd patient data ---
    print("\n[1/4] Parsing Elo-KRd Excel...")
    elo_df = parse_elo_krd()
    elo_path = os.path.join(DATA_DIR, "elo_krd_patients.csv")
    elo_df.to_csv(elo_path, index=False)
    n_patients = len(elo_df)
    n_evaluable = elo_df["met_primary_endpoint_window"].sum()
    n_responders = elo_df.loc[
        elo_df["met_primary_endpoint_window"] == 1, "primary_endpoint"
    ].sum()
    print(f"  Patients: {n_patients}")
    print(f"  Evaluable (met endpoint window): {n_evaluable}")
    print(f"  Responders (sCR/MRD-neg at C8): {n_responders}")
    print(f"  Rate among evaluable: {n_responders}/{n_evaluable} = {n_responders/n_evaluable:.1%}")
    assert n_patients == 30, f"Expected 30 patients, got {n_patients}"
    assert len(elo_df["subject_id"].unique()) == n_patients, "Duplicate subject IDs"

    # --- 2. Trials data ---
    print("\n[2/4] Building trials_data.csv...")
    trials_df = build_trials_data()
    trials_path = os.path.join(DATA_DIR, "trials_data.csv")
    trials_df.to_csv(trials_path, index=False)
    n_arms = len(trials_df)
    print(f"  Trial arms: {n_arms}")
    print(f"  Unique trials: {trials_df['trial_name'].nunique()}")
    assert all(trials_df["p_hat"].between(0, 1)), "p_hat out of range"
    assert all(trials_df["n"] > 0), "n must be positive"
    n_1e5 = (trials_df["mrd_threshold"] == "1e-5").sum()
    print(f"  Arms at 10^-5: {n_1e5}, at 10^-4: {n_arms - n_1e5}")

    # --- 3. Drug-target map ---
    print("\n[3/4] Building drug_target_map.csv...")
    drug_df = build_drug_target_map()
    drug_path = os.path.join(DATA_DIR, "drug_target_map.csv")
    drug_df.to_csv(drug_path, index=False)
    print(f"  Drugs mapped: {len(drug_df)}")

    # Validate: all drugs in trials_data appear in drug_target_map
    trial_drugs = set()
    for d in trials_df["drugs"]:
        trial_drugs.update(d.split(","))
    mapped_drugs = set(drug_df["drug_name"])
    missing = trial_drugs - mapped_drugs
    if missing:
        warnings.warn(f"Drugs in trials_data but not in drug_target_map: {missing}")
    else:
        print("  All trial drugs mapped.")

    # --- 4. Simulated patient data ---
    print("\n[4/4] Simulating patient data for MASTER and MANHATTAN...")

    master_df = simulate_patient_responses(
        n=123, p_hat=0.71, frac_high_risk=0.32,
        p_hat_hr=0.63, p_hat_sr=0.75,
        frac_iss3=0.13, median_age=57, seed=2,
    )
    master_path = os.path.join(DATA_DIR, "master_simulated_patients.csv")
    master_df.to_csv(master_path, index=False)
    print(f"  MASTER: n={len(master_df)}, response rate={master_df['response'].mean():.3f} (target 0.71)")

    manhattan_df = simulate_patient_responses(
        n=41, p_hat=0.71, frac_high_risk=0.41,
        p_hat_hr=0.65, p_hat_sr=0.79,
        frac_iss3=0.15, median_age=58, seed=123,
    )
    manhattan_path = os.path.join(DATA_DIR, "manhattan_simulated_patients.csv")
    manhattan_df.to_csv(manhattan_path, index=False)
    print(f"  MANHATTAN: n={len(manhattan_df)}, response rate={manhattan_df['response'].mean():.3f} (target 0.71)")

    print("\n" + "=" * 60)
    print("All data files written to data/")
    print("=" * 60)


if __name__ == "__main__":
    main()
